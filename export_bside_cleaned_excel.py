from __future__ import annotations

import hashlib
from collections import Counter
from datetime import timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill

from generate_person_session_dashboard import (
    CLARITY_SESSION_GAP_SECONDS,
    OUTPUT_DIR,
    PATH_CONTAINS_RULES,
    PATH_SUFFIX_RULES,
    SKIP_FIRST_DAYS,
    TARGET_METHOD,
    browser_family,
    build_sessions,
    device_type,
    load_bside_records,
    os_family,
)


SPREADSHEET_DIR = OUTPUT_DIR / "spreadsheet"
SESSION_GAP_SECONDS = CLARITY_SESSION_GAP_SECONDS


HEADER_FILL = PatternFill("solid", fgColor="DCEBFF")
SUBHEADER_FILL = PatternFill("solid", fgColor="EAF7F4")
TITLE_FONT = Font(bold=True, size=12, color="0F172A")
HEADER_FONT = Font(bold=True, color="0F172A")
NORMAL_FONT = Font(color="334155")


def data_dictionary_rows() -> list[list[str]]:
    return [
        ["口径说明", "项目", "说明项名称", "导出脚本固定枚举项", "由 write_scope_sheet 按固定顺序写入", "用于描述数据范围、过滤规则、统计结果和排除项"],
        ["口径说明", "值", "对应项目的值", "脚本常量、清洗结果、会话结果", "来自 SESSION_GAP_SECONDS、SKIP_FIRST_DAYS、TARGET_METHOD、len(records)、len(sessions)、sum(dropped.values()) 等", "数值和文本会混合出现"],
        ["口径说明", "说明", "对项目和值的解释", "导出脚本固定说明文本", "由 write_scope_sheet 写死", "用于对外说明口径"],
        ["每日人次", "date", "统计日期", "原始日志字段 dt", "dt 解析后格式化为 YYYY-MM-DD，并按会话开始时间所属日期统计", "每日一行"],
        ["每日人次", "estimated_sessions", "当日估算人次", "人次明细 sheet", "对同一天的 session_id 计数", "口径为 Clarity 风格 30 分钟无活动切新会话"],
        ["入口页汇总", "entry_page", "会话首个页面路径", "原始日志字段 uri", "uri 经 normalize_page 和异常路径过滤后，取每个会话的首个 page", "按人次首访页汇总，不是 PV"],
        ["入口页汇总", "sessions", "该入口页对应的人次", "人次明细 sheet", "对 entry_page 相同的会话计数并降序排序", "同一会话只记 1 次"],
        ["设备浏览器汇总", "browser", "浏览器类别", "原始日志字段 ua", "用 browser_family(user_agent) 归类为 Chrome/Edge/Firefox/Safari/WeChat/Other", "A-B 列为浏览器区块"],
        ["设备浏览器汇总", "sessions", "浏览器类别对应人次", "人次明细 sheet", "对 session.browser 计数并降序排序", "A-B 列中的 sessions 对应 browser"],
        ["设备浏览器汇总", "device_type", "设备类型", "原始日志字段 ua", "用 device_type(user_agent) 归类为 Desktop/Mobile/Tablet", "D-E 列为设备区块"],
        ["设备浏览器汇总", "sessions", "设备类型对应人次", "人次明细 sheet", "对 session.device_type 计数并降序排序", "D-E 列中的 sessions 对应 device_type"],
        ["设备浏览器汇总", "os", "操作系统类别", "原始日志字段 ua", "用 os_family(user_agent) 归类为 Windows/macOS/Android/iOS/Linux/Other", "G-H 列为 OS 区块"],
        ["设备浏览器汇总", "sessions", "操作系统对应人次", "人次明细 sheet", "对 session.os 计数并降序排序", "G-H 列中的 sessions 对应 os"],
        ["指纹字典", "fingerprint", "浏览器指纹近似值", "原始日志字段 ua", "对完整 User-Agent 做 SHA1 哈希并截取前 12 位", "当前 workbook 的核心去重键"],
        ["指纹字典", "requests", "该指纹对应的清洗后请求数", "清洗后请求明细 sheet", "对 fingerprint 相同的请求计数", "不是会话数"],
        ["指纹字典", "browser", "该指纹样本所属浏览器分类", "原始日志字段 ua", "对该 fingerprint 首次出现的 user_agent 调用 browser_family", "样本字段"],
        ["指纹字典", "os", "该指纹样本所属操作系统分类", "原始日志字段 ua", "对该 fingerprint 首次出现的 user_agent 调用 os_family", "样本字段"],
        ["指纹字典", "device_type", "该指纹样本所属设备分类", "原始日志字段 ua", "对该 fingerprint 首次出现的 user_agent 调用 device_type", "样本字段"],
        ["指纹字典", "sample_proxy_ip", "样本代理 IP", "原始日志字段 ip", "取该 fingerprint 首次出现记录的 proxy_ip", "仅用于排查，不用于会话键"],
        ["指纹字典", "sample_user_agent", "样本 User-Agent 原文", "原始日志字段 ua", "取该 fingerprint 首次出现记录的完整 ua", "用于回查指纹含义"],
        ["清洗后请求明细", "request_time", "请求时间", "原始日志字段 dt", "解析为本地无时区 datetime 后以 YYYY-MM-DD HH:MM:SS 输出", "每条合格请求一行"],
        ["清洗后请求明细", "date", "请求日期", "原始日志字段 dt", "由 request_time 截取到日", "便于按天筛选"],
        ["清洗后请求明细", "proxy_ip", "日志中的代理 IP", "原始日志字段 ip", "直接写入原始日志中的 ip", "B-side 下通常为内网代理 IP"],
        ["清洗后请求明细", "fingerprint", "浏览器指纹近似值", "原始日志字段 ua", "对完整 ua 做 SHA1 哈希并截取前 12 位", "与人次明细关联键之一"],
        ["清洗后请求明细", "browser", "浏览器类别", "原始日志字段 ua", "用 browser_family(user_agent) 归类", "派生字段"],
        ["清洗后请求明细", "os", "操作系统类别", "原始日志字段 ua", "用 os_family(user_agent) 归类", "派生字段"],
        ["清洗后请求明细", "device_type", "设备类别", "原始日志字段 ua", "用 device_type(user_agent) 归类", "派生字段"],
        ["清洗后请求明细", "page", "清洗后的页面路径", "原始日志字段 uri", "uri 经 normalize_page 后，再经过异常路径排除规则", "当前分析的核心页面字段"],
        ["清洗后请求明细", "referer", "来源页", "原始日志字段 referer", "原值为空或 - 时替换为 (direct / none)", "用于分析来源渠道与入口来源"],
        ["清洗后请求明细", "user_agent", "原始浏览器标识", "原始日志字段 ua", "直接写入原始 ua", "用于回查分类和指纹"],
        ["人次明细", "session_id", "会话唯一编号", "build_sessions 生成结果", "按生成顺序从 1 开始递增", "每条会话一行"],
        ["人次明细", "day", "会话开始日期", "会话 start", "取 session.start 所属日期", "用于按天统计人次"],
        ["人次明细", "start", "会话开始时间", "清洗后请求明细 sheet", "取该会话第一条请求的 request_time", "会话起点"],
        ["人次明细", "end", "会话结束时间", "清洗后请求明细 sheet", "取该会话最后一条请求的 request_time", "会话终点"],
        ["人次明细", "proxy_ip", "会话首条请求的代理 IP", "原始日志字段 ip", "取创建会话时的 proxy_ip", "仅保留用于排查"],
        ["人次明细", "fingerprint", "会话主键使用的浏览器指纹近似值", "原始日志字段 ua", "对完整 ua 做 SHA1 哈希并截取前 12 位", "会话聚合键"],
        ["人次明细", "browser", "会话浏览器类别", "原始日志字段 ua", "按创建会话时的 ua 调用 browser_family", "派生字段"],
        ["人次明细", "os", "会话操作系统类别", "原始日志字段 ua", "按创建会话时的 ua 调用 os_family", "派生字段"],
        ["人次明细", "device_type", "会话设备类别", "原始日志字段 ua", "按创建会话时的 ua 调用 device_type", "派生字段"],
        ["人次明细", "entry_page", "会话入口页", "清洗后请求明细 sheet", "取该会话第一条请求的 page", "用于入口页分析"],
        ["人次明细", "referer", "会话首条请求来源页", "原始日志字段 referer", "取创建会话时的 referer", "用于近似分析来源"],
        ["人次明细", "pageviews", "会话内页面请求数", "清洗后请求明细 sheet", "同一 session_id 下累计请求条数", "仅统计清洗后的合格页面请求"],
        ["人次明细", "duration_seconds", "会话持续秒数", "会话 start/end", "end - start 的秒数，最小为 0", "没有前端事件时长，因此是日志请求时长"],
    ]


def header_row(ws, values: list[str]) -> None:
    row = []
    for value in values:
        cell = WriteOnlyCell(ws, value=value)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        row.append(cell)
    ws.append(row)


def normal_row(ws, values: list[object]) -> None:
    row = []
    for value in values:
        cell = WriteOnlyCell(ws, value=value)
        cell.font = NORMAL_FONT
        cell.alignment = Alignment(vertical="top")
        row.append(cell)
    ws.append(row)


def write_data_dictionary_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("字段字典")
    ws.freeze_panes = "A2"
    widths = {
        "A": 18,
        "B": 18,
        "C": 28,
        "D": 26,
        "E": 56,
        "F": 30,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    header_row(ws, ["sheet_name", "field_name", "业务含义", "字段来源", "获取方式 / 计算方式", "备注"])
    for row in data_dictionary_rows():
        normal_row(ws, row)


def write_scope_sheet(
    wb: Workbook,
    records: list[tuple],
    sessions: list,
    dropped: Counter,
    output_path: Path,
) -> None:
    ws = wb.create_sheet("口径说明")
    ws.freeze_panes = "A5"
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 20

    title = WriteOnlyCell(ws, value="B-side 清洗后数据导出")
    title.font = TITLE_FONT
    title.fill = SUBHEADER_FILL
    ws.append([title, "", ""])

    normal_row(ws, ["导出文件", output_path.name, ""])
    normal_row(ws, ["生成位置", str(output_path), ""])
    ws.append([])

    header_row(ws, ["项目", "值", "说明"])
    normal_row(ws, ["数据范围", "仅 B-side", "不含 C-side"])
    normal_row(ws, ["人次会话阈值", SESSION_GAP_SECONDS, "与 Clarity 对齐：同一浏览器指纹 30 分钟无活动后切新会话"])
    normal_row(ws, ["会话键", "完整 UA 哈希", "B-side 代理内网 IP 不参与会话键"])
    normal_row(ws, ["保留方法", TARGET_METHOD, "只保留 GET"])
    normal_row(ws, ["跳过最早天数", SKIP_FIRST_DAYS, "最早 3 天数据不完整，整体剔除"])
    normal_row(ws, ["过滤静态资源", "是", "全局过滤 /wp-content、/wp-includes 及常见资源扩展名"])
    normal_row(ws, ["剔除爬虫 UA", "是", "按 UA 规则识别 bot / spider / crawler 等"])
    normal_row(ws, ["剔除关键字路径", "是", "删除包含 " + ", ".join(token for token, _ in PATH_CONTAINS_RULES) + " 的路径"])
    normal_row(ws, ["剔除残缺路径", "是", "删除以 " + ", ".join(PATH_SUFFIX_RULES[idx][0] for idx in range(len(PATH_SUFFIX_RULES))) + " 结尾的路径"])
    normal_row(ws, ["剔除隐藏探测路径", "是", "例如 /.env、/.git/config"])
    normal_row(ws, ["清洗后页面请求数", len(records), "进入人次估算前的合格页面请求"])
    normal_row(ws, ["估算人次", len(sessions), "按浏览器指纹近似还原 Clarity 会话后的结果"])
    normal_row(ws, ["剔除记录数", sum(dropped.values()), "各类排除项之和"])
    ws.append([])

    header_row(ws, ["排除项", "数量", ""])
    for key, value in dropped.items():
        normal_row(ws, [key, value, ""])


def write_summary_sheets(wb: Workbook, records: list[tuple], sessions: list) -> None:
    daily_sessions = Counter(session.day for session in sessions)
    entry_pages = Counter(session.entry_page for session in sessions)
    browser_counter = Counter(session.browser for session in sessions)
    device_counter = Counter(session.device_type for session in sessions)
    os_counter = Counter(session.os for session in sessions)

    ws_daily = wb.create_sheet("每日人次")
    ws_daily.freeze_panes = "A2"
    ws_daily.column_dimensions["A"].width = 14
    ws_daily.column_dimensions["B"].width = 14
    header_row(ws_daily, ["date", "estimated_sessions"])
    for day in sorted(daily_sessions):
        normal_row(ws_daily, [day, daily_sessions[day]])

    ws_entry = wb.create_sheet("入口页汇总")
    ws_entry.freeze_panes = "A2"
    ws_entry.column_dimensions["A"].width = 40
    ws_entry.column_dimensions["B"].width = 14
    header_row(ws_entry, ["entry_page", "sessions"])
    for page, count in entry_pages.most_common():
        normal_row(ws_entry, [page, count])

    ws_browser = wb.create_sheet("设备浏览器汇总")
    ws_browser.freeze_panes = "A2"
    ws_browser.column_dimensions["A"].width = 18
    ws_browser.column_dimensions["B"].width = 12
    ws_browser.column_dimensions["D"].width = 18
    ws_browser.column_dimensions["E"].width = 12
    ws_browser.column_dimensions["G"].width = 18
    ws_browser.column_dimensions["H"].width = 12
    header_row(ws_browser, ["browser", "sessions", "", "device_type", "sessions", "", "os", "sessions"])

    max_len = max(len(browser_counter), len(device_counter), len(os_counter))
    browser_rows = browser_counter.most_common()
    device_rows = device_counter.most_common()
    os_rows = os_counter.most_common()
    for idx in range(max_len):
        left = list(browser_rows[idx]) if idx < len(browser_rows) else ["", ""]
        mid = list(device_rows[idx]) if idx < len(device_rows) else ["", ""]
        right = list(os_rows[idx]) if idx < len(os_rows) else ["", ""]
        normal_row(ws_browser, left + [""] + mid + [""] + right)

    fp_counter = Counter()
    fp_pages = Counter()
    fp_sample = {}
    for _, _, proxy_ip, user_agent, page, _ in records:
        fingerprint = hashlib.sha1(user_agent.encode("utf-8", "ignore")).hexdigest()[:12]
        fp_counter[fingerprint] += 1
        fp_pages[(fingerprint, page)] += 1
        fp_sample.setdefault(
            fingerprint,
            {
                "browser": browser_family(user_agent),
                "os": os_family(user_agent),
                "device_type": device_type(user_agent),
                "proxy_ip": proxy_ip,
                "user_agent": user_agent,
            },
        )

    ws_fp = wb.create_sheet("指纹字典")
    ws_fp.freeze_panes = "A2"
    for col in ("A", "B", "C", "D", "E", "F", "G"):
        ws_fp.column_dimensions[col].width = 18
    ws_fp.column_dimensions["G"].width = 90
    header_row(ws_fp, ["fingerprint", "requests", "browser", "os", "device_type", "sample_proxy_ip", "sample_user_agent"])
    for fingerprint, count in fp_counter.most_common():
        sample = fp_sample[fingerprint]
        normal_row(
            ws_fp,
            [
                fingerprint,
                count,
                sample["browser"],
                sample["os"],
                sample["device_type"],
                sample["proxy_ip"],
                sample["user_agent"],
            ],
        )


def write_requests_sheet(wb: Workbook, records: list[tuple]) -> None:
    ws = wb.create_sheet("清洗后请求明细")
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 40
    ws.column_dimensions["I"].width = 60
    ws.column_dimensions["J"].width = 90

    header_row(
        ws,
        [
            "request_time",
            "date",
            "proxy_ip",
            "fingerprint",
            "browser",
            "os",
            "device_type",
            "page",
            "referer",
            "user_agent",
        ],
    )
    for dt, day, proxy_ip, user_agent, page, referer in records:
        fingerprint = hashlib.sha1(user_agent.encode("utf-8", "ignore")).hexdigest()[:12]
        normal_row(
            ws,
            [
                dt.isoformat(sep=" "),
                day,
                proxy_ip,
                fingerprint,
                browser_family(user_agent),
                os_family(user_agent),
                device_type(user_agent),
                page,
                referer,
                user_agent,
            ],
        )


def write_sessions_sheet(wb: Workbook, sessions: list) -> None:
    ws = wb.create_sheet("人次明细")
    ws.freeze_panes = "A2"
    for col in ("A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M"):
        ws.column_dimensions[col].width = 18
    ws.column_dimensions["J"].width = 36
    ws.column_dimensions["K"].width = 60
    header_row(
        ws,
        [
            "session_id",
            "day",
            "start",
            "end",
            "proxy_ip",
            "fingerprint",
            "browser",
            "os",
            "device_type",
            "entry_page",
            "referer",
            "pageviews",
            "duration_seconds",
        ],
    )
    for session in sessions:
        duration_seconds = max(0, int((session.end - session.start).total_seconds()))
        normal_row(
            ws,
            [
                session.session_id,
                session.day,
                session.start.isoformat(sep=" "),
                session.end.isoformat(sep=" "),
                session.proxy_ip,
                session.fingerprint,
                session.browser,
                session.os,
                session.device_type,
                session.entry_page,
                session.referer,
                session.pageviews,
                duration_seconds,
            ],
        )


def main() -> None:
    records, dropped = load_bside_records()
    sessions = build_sessions(records, timedelta(seconds=SESSION_GAP_SECONDS))

    SPREADSHEET_DIR.mkdir(parents=True, exist_ok=True)
    output_path = SPREADSHEET_DIR / f"bside_cleaned_data_clarity_{SESSION_GAP_SECONDS}s.xlsx"

    wb = Workbook(write_only=True)
    write_data_dictionary_sheet(wb)
    write_scope_sheet(wb, records, sessions, dropped, output_path)
    write_summary_sheets(wb, records, sessions)
    write_requests_sheet(wb, records)
    write_sessions_sheet(wb, sessions)
    wb.save(output_path)

    print(output_path)


if __name__ == "__main__":
    main()
