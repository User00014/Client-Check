from __future__ import annotations

from collections import Counter
from datetime import timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill

from generate_bside_bot_dashboard import (
    BOT_SESSION_GAP_SECONDS,
    OUTPUT_DIR,
    build_bot_sessions,
    bot_group,
    load_bside_bot_records,
)


SPREADSHEET_DIR = OUTPUT_DIR / "spreadsheet"
HEADER_FILL = PatternFill("solid", fgColor="DCEBFF")
SUBHEADER_FILL = PatternFill("solid", fgColor="EAF7F4")
TITLE_FONT = Font(bold=True, size=12, color="0F172A")
HEADER_FONT = Font(bold=True, color="0F172A")
NORMAL_FONT = Font(color="334155")


def header_row(ws, values: list[str]) -> None:
    row = []
    for value in values:
        cell = WriteOnlyCell(ws, value=value)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        row.append(cell)
    ws.append(row)


def normal_row(ws, values: list[object]) -> None:
    row = []
    for value in values:
        cell = WriteOnlyCell(ws, value=value)
        cell.font = NORMAL_FONT
        cell.alignment = Alignment(vertical="top")
        row.append(cell)
    ws.append(row)


def data_dictionary_rows() -> list[list[str]]:
    return [
        ["字段字典", "sheet_name", "字段所属工作表", "导出脚本固定配置", "由 data_dictionary_rows 枚举生成", "总览字段"],
        ["字段字典", "field_name", "字段名", "导出脚本固定配置", "由 data_dictionary_rows 枚举生成", "总览字段"],
        ["字段字典", "业务含义", "字段在业务上的解释", "导出脚本固定说明", "人工整理写入", ""],
        ["字段字典", "字段来源", "字段直接来自原始日志或派生结果", "脚本逻辑", "人工整理写入", ""],
        ["字段字典", "获取方式 / 计算方式", "字段如何从原始日志得到", "脚本逻辑", "人工整理写入", ""],
        ["字段字典", "备注", "补充说明", "导出脚本固定说明", "人工整理写入", ""],
        ["口径说明", "项目", "说明项名称", "导出脚本固定枚举项", "由 write_scope_sheet 按固定顺序写入", "用于说明范围、过滤规则和统计值"],
        ["口径说明", "值", "说明项对应的值", "脚本常量或统计结果", "由常量和聚合结果写入", ""],
        ["口径说明", "说明", "对项目和值的解释", "导出脚本固定说明", "由 write_scope_sheet 写入", ""],
        ["每日Bot请求", "date", "统计日期", "原始日志字段 dt", "按清洗后 bot 请求所属日期汇总", ""],
        ["每日Bot请求", "bot_requests", "当日 bot 请求量", "Bot清洗后请求明细", "对同一天的清洗后 bot 请求计数", ""],
        ["每日Bot会话", "date", "统计日期", "会话 start", "按 bot session 开始时间所属日期汇总", ""],
        ["每日Bot会话", "bot_sessions", "当日 bot session 数", "Bot会话明细", "对同一天的 session_id 计数", "会话阈值 1 分钟"],
        ["分类汇总", "bot_category", "bot 分类", "原始日志字段 ua", "按 classify_bot 规则映射得到", "主分类字段"],
        ["分类汇总", "bot_group", "bot 分组", "bot_category", "AI Retrieval/Training/Indexer 归入 AI Bot，SEO Bot 单独保留，其余归 Other External Bot", "汇总字段"],
        ["分类汇总", "requests", "分类对应请求量", "Bot清洗后请求明细", "对 bot_category 计数", ""],
        ["分类汇总", "sessions", "分类对应 session 数", "Bot会话明细", "对 bot_category 计数", ""],
        ["家族汇总", "bot_family", "bot 家族名称", "原始日志字段 ua", "按 classify_bot 规则映射得到", "命名粒度比分类更细"],
        ["家族汇总", "bot_category", "家族所属分类", "bot_family", "从家族映射到分类", ""],
        ["家族汇总", "bot_group", "家族所属分组", "bot_category", "由 bot_group(category) 计算", ""],
        ["家族汇总", "requests", "家族对应请求量", "Bot清洗后请求明细", "对 bot_family 计数", ""],
        ["家族汇总", "sessions", "家族对应 session 数", "Bot会话明细", "对 bot_family 计数", ""],
        ["入口页汇总", "entry_page", "bot session 首个页面", "原始日志字段 uri", "uri 经 normalize_page 和异常路径过滤后，取每个 session 的首个 page", "不是 PV"],
        ["入口页汇总", "sessions", "以该页面为入口的 bot session 数", "Bot会话明细", "对 entry_page 相同的 session 计数", ""],
        ["Bot清洗后请求明细", "request_time", "请求时间", "原始日志字段 dt", "解析为 datetime 后输出", "每条清洗后 bot 请求一行"],
        ["Bot清洗后请求明细", "date", "请求日期", "request_time", "格式化为 YYYY-MM-DD", ""],
        ["Bot清洗后请求明细", "proxy_ip", "日志中的代理 IP", "原始日志字段 ip", "直接写入", "B-side 下通常为内网代理 IP"],
        ["Bot清洗后请求明细", "fingerprint", "bot 指纹近似值", "原始日志字段 ua", "对完整 ua 做 SHA1 哈希并截取前 12 位", "也是 bot session 键"],
        ["Bot清洗后请求明细", "bot_category", "bot 分类", "原始日志字段 ua", "调用 classify_bot 映射得到", ""],
        ["Bot清洗后请求明细", "bot_group", "bot 分组", "bot_category", "调用 bot_group 映射得到", ""],
        ["Bot清洗后请求明细", "bot_family", "bot 家族", "原始日志字段 ua", "调用 classify_bot 映射得到", ""],
        ["Bot清洗后请求明细", "page", "清洗后的页面路径", "原始日志字段 uri", "uri 经 normalize_page 与异常路径过滤后保留", ""],
        ["Bot清洗后请求明细", "referer", "来源页", "原始日志字段 referer", "空值或 - 统一写为 (direct / none)", ""],
        ["Bot清洗后请求明细", "user_agent", "原始 User-Agent", "原始日志字段 ua", "直接写入", "用于回查 bot 家族"],
        ["Bot会话明细", "session_id", "bot session 编号", "会话聚合结果", "按生成顺序从 1 开始递增", "每条 bot session 一行"],
        ["Bot会话明细", "day", "session 所属日期", "session start", "取 start 所属日期", ""],
        ["Bot会话明细", "start", "session 开始时间", "Bot清洗后请求明细", "该 session 第一条请求的 request_time", ""],
        ["Bot会话明细", "end", "session 结束时间", "Bot清洗后请求明细", "该 session 最后一条请求的 request_time", ""],
        ["Bot会话明细", "proxy_ip", "session 首条请求的代理 IP", "原始日志字段 ip", "创建 session 时记录的 proxy_ip", "仅用于排查"],
        ["Bot会话明细", "fingerprint", "bot session 键", "原始日志字段 ua", "对完整 ua 做 SHA1 哈希并截取前 12 位", ""],
        ["Bot会话明细", "bot_category", "session 的 bot 分类", "原始日志字段 ua", "会话首条请求分类", ""],
        ["Bot会话明细", "bot_group", "session 的 bot 分组", "bot_category", "调用 bot_group 映射得到", ""],
        ["Bot会话明细", "bot_family", "session 的 bot 家族", "原始日志字段 ua", "会话首条请求家族", ""],
        ["Bot会话明细", "entry_page", "session 首访页面", "Bot清洗后请求明细", "该 session 第一条请求的 page", ""],
        ["Bot会话明细", "referer", "session 首访来源页", "原始日志字段 referer", "会话首条请求的 referer", ""],
        ["Bot会话明细", "pageviews", "session 内页面请求数", "Bot清洗后请求明细", "同一 session_id 下累计请求条数", ""],
        ["Bot会话明细", "duration_seconds", "session 持续秒数", "session start/end", "end - start 的秒数，最小为 0", "是日志跨度，不是前端停留时长"],
    ]


def write_data_dictionary_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("字段字典")
    ws.freeze_panes = "A2"
    for col, width in {"A": 18, "B": 18, "C": 28, "D": 26, "E": 56, "F": 24}.items():
        ws.column_dimensions[col].width = width
    header_row(ws, ["sheet_name", "field_name", "业务含义", "字段来源", "获取方式 / 计算方式", "备注"])
    for row in data_dictionary_rows():
        normal_row(ws, row)


def write_scope_sheet(wb: Workbook, records, sessions, dropped: Counter, output_path: Path) -> None:
    ws = wb.create_sheet("口径说明")
    ws.freeze_panes = "A5"
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 28

    title = WriteOnlyCell(ws, value="B-side Bot 清洗后数据导出")
    title.font = TITLE_FONT
    title.fill = SUBHEADER_FILL
    ws.append([title, "", ""])

    normal_row(ws, ["导出文件", output_path.name, ""])
    normal_row(ws, ["生成位置", str(output_path), ""])
    ws.append([])

    header_row(ws, ["项目", "值", "说明"])
    normal_row(ws, ["数据范围", "仅 B-side", "不含 C-side 和真实用户请求"])
    normal_row(ws, ["Bot session 阈值", BOT_SESSION_GAP_SECONDS, "同一 bot 指纹 1 分钟无活动后切新 session"])
    normal_row(ws, ["Session 键", "完整 UA 哈希", "B-side 代理内网 IP 不参与 bot session 键"])
    normal_row(ws, ["保留方法", "GET", "只保留 GET 页面请求"])
    normal_row(ws, ["跳过最早天数", 3, "最早 3 天数据不完整，整体剔除"])
    normal_row(ws, ["过滤静态资源", "是", "全局过滤 /wp-content、/wp-includes 及常见静态资源扩展名"])
    normal_row(ws, ["保留 bot 类型", "客户相关外部 bot", "保留 AI Retrieval / Training / Indexer、SEO Bot、Social / Platform Bot、Other External Bot"])
    normal_row(ws, ["排除 bot 类型", "是", "排除 Script / Client Bot 和 Scanner / Monitor Bot"])
    normal_row(ws, ["路径清洗", "是", "删除 admin/wp-/php/tel/feed/ashx/backup、残缺路径和隐藏探测路径"])
    normal_row(ws, ["清洗后 bot 请求数", len(records), "进入 bot session 统计前的合格 bot 页面请求"])
    normal_row(ws, ["Bot sessions", len(sessions), "按 1 分钟无活动规则聚合后的 bot session"])
    normal_row(ws, ["剔除记录数", sum(dropped.values()), "各类排除项之和"])
    ws.append([])

    header_row(ws, ["排除项", "数量", ""])
    for key, value in dropped.items():
        normal_row(ws, [key, value, ""])


def write_daily_sheets(wb: Workbook, records, sessions) -> None:
    daily_requests = Counter(record.day for record in records)
    daily_sessions = Counter(session.day for session in sessions)

    ws_req = wb.create_sheet("每日Bot请求")
    ws_req.freeze_panes = "A2"
    ws_req.column_dimensions["A"].width = 14
    ws_req.column_dimensions["B"].width = 14
    header_row(ws_req, ["date", "bot_requests"])
    for day in sorted(daily_requests):
        normal_row(ws_req, [day, daily_requests[day]])

    ws_sess = wb.create_sheet("每日Bot会话")
    ws_sess.freeze_panes = "A2"
    ws_sess.column_dimensions["A"].width = 14
    ws_sess.column_dimensions["B"].width = 14
    header_row(ws_sess, ["date", "bot_sessions"])
    for day in sorted(daily_sessions):
        normal_row(ws_sess, [day, daily_sessions[day]])


def write_summary_sheets(wb: Workbook, records, sessions) -> None:
    category_requests = Counter(record.bot_category for record in records)
    category_sessions = Counter(session.bot_category for session in sessions)
    family_requests = Counter(record.bot_family for record in records)
    family_sessions = Counter(session.bot_family for session in sessions)
    family_to_category = {}
    for record in records:
        family_to_category.setdefault(record.bot_family, record.bot_category)
    entry_sessions = Counter(session.entry_page for session in sessions)

    ws_cat = wb.create_sheet("分类汇总")
    ws_cat.freeze_panes = "A2"
    ws_cat.column_dimensions["A"].width = 22
    ws_cat.column_dimensions["B"].width = 20
    ws_cat.column_dimensions["C"].width = 14
    ws_cat.column_dimensions["D"].width = 14
    header_row(ws_cat, ["bot_category", "bot_group", "requests", "sessions"])
    for category, requests in category_requests.most_common():
        normal_row(ws_cat, [category, bot_group(category), requests, category_sessions.get(category, 0)])

    ws_family = wb.create_sheet("家族汇总")
    ws_family.freeze_panes = "A2"
    ws_family.column_dimensions["A"].width = 22
    ws_family.column_dimensions["B"].width = 22
    ws_family.column_dimensions["C"].width = 20
    ws_family.column_dimensions["D"].width = 14
    ws_family.column_dimensions["E"].width = 14
    header_row(ws_family, ["bot_family", "bot_category", "bot_group", "requests", "sessions"])
    for family, requests in family_requests.most_common():
        category = family_to_category.get(family, "Other External Bot")
        normal_row(ws_family, [family, category, bot_group(category), requests, family_sessions.get(family, 0)])

    ws_entry = wb.create_sheet("入口页汇总")
    ws_entry.freeze_panes = "A2"
    ws_entry.column_dimensions["A"].width = 40
    ws_entry.column_dimensions["B"].width = 14
    header_row(ws_entry, ["entry_page", "sessions"])
    for page, count in entry_sessions.most_common():
        normal_row(ws_entry, [page, count])


def write_requests_sheet(wb: Workbook, records) -> None:
    ws = wb.create_sheet("Bot清洗后请求明细")
    ws.freeze_panes = "A2"
    widths = {"A": 22, "B": 14, "C": 16, "D": 16, "E": 20, "F": 18, "G": 18, "H": 40, "I": 60, "J": 90}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    header_row(
        ws,
        ["request_time", "date", "proxy_ip", "fingerprint", "bot_category", "bot_group", "bot_family", "page", "referer", "user_agent"],
    )
    for record in records:
        normal_row(
            ws,
            [
                record.request_time.isoformat(sep=" "),
                record.day,
                record.proxy_ip,
                record.fingerprint,
                record.bot_category,
                bot_group(record.bot_category),
                record.bot_family,
                record.page,
                record.referer,
                record.user_agent,
            ],
        )


def write_sessions_sheet(wb: Workbook, sessions) -> None:
    ws = wb.create_sheet("Bot会话明细")
    ws.freeze_panes = "A2"
    widths = {"A": 12, "B": 14, "C": 22, "D": 22, "E": 16, "F": 16, "G": 20, "H": 18, "I": 18, "J": 36, "K": 60, "L": 12, "M": 16}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    header_row(
        ws,
        ["session_id", "day", "start", "end", "proxy_ip", "fingerprint", "bot_category", "bot_group", "bot_family", "entry_page", "referer", "pageviews", "duration_seconds"],
    )
    for session in sessions:
        normal_row(
            ws,
            [
                session.session_id,
                session.day,
                session.start.isoformat(sep=" "),
                session.end.isoformat(sep=" "),
                session.proxy_ip,
                session.fingerprint,
                session.bot_category,
                bot_group(session.bot_category),
                session.bot_family,
                session.entry_page,
                session.referer,
                session.pageviews,
                max(0, int((session.end - session.start).total_seconds())),
            ],
        )


def main() -> None:
    records, dropped = load_bside_bot_records()
    sessions = build_bot_sessions(records, timedelta(seconds=BOT_SESSION_GAP_SECONDS))

    SPREADSHEET_DIR.mkdir(parents=True, exist_ok=True)
    output_path = SPREADSHEET_DIR / f"bside_bot_cleaned_data_{BOT_SESSION_GAP_SECONDS}s.xlsx"

    wb = Workbook(write_only=True)
    write_data_dictionary_sheet(wb)
    write_scope_sheet(wb, records, sessions, dropped, output_path)
    write_daily_sheets(wb, records, sessions)
    write_summary_sheets(wb, records, sessions)
    write_requests_sheet(wb, records)
    write_sessions_sheet(wb, sessions)
    wb.save(output_path)
    print(output_path)


if __name__ == "__main__":
    main()
